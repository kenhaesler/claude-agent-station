"""Vision attachment helpers: sanitise filenames, sniff MIME, extract text.

Spec: docs/superpowers/specs/2026-05-21-vision-reference-files-design.md.
"""
from __future__ import annotations

import base64
import io
import os
import re
import shutil
import uuid as _uuid
from datetime import datetime, timezone
from pathlib import Path

import magic
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings as _settings
from app.models import VisionChatAttachment

# Anthropic's forbidden-character set plus path separators
_FORBIDDEN = re.compile(r'[<>:"|?*\\/]')


def sanitize_filename(raw: str) -> str:
    """Strip path components and forbidden characters; cap length at 255.

    Raises ValueError if nothing usable remains.
    """
    base = os.path.basename(raw.replace("\\", "/"))
    cleaned = _FORBIDDEN.sub("", base)
    cleaned = cleaned.lstrip(". ").rstrip(" ")
    if not cleaned:
        raise ValueError("filename empty after sanitisation")
    if len(cleaned) > 255:
        root, ext = os.path.splitext(cleaned)
        keep = 255 - len(ext)
        cleaned = root[:keep] + ext
    return cleaned


class AttachmentRejected(Exception):
    """Raised when an upload fails validation (size, MIME, or sanitisation)."""


ALLOWED_MIMES = frozenset({
    "application/pdf",
    "image/png", "image/jpeg", "image/gif", "image/webp",
    "text/plain", "text/markdown", "text/csv",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # xlsx
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",  # docx
})

# Extension → expected MIME, used to normalise libmagic quirks (e.g. csv often
# detected as text/plain). The extension is advisory only; the sniffed type
# wins for security-sensitive checks like rejecting zips disguised as docx.
_EXT_HINTS: dict[str, str] = {
    ".csv": "text/csv",
    ".md": "text/markdown",
    ".txt": "text/plain",
}


def sniff_and_validate_mime(data: bytes, *, declared_filename: str) -> str:
    """Sniff the MIME of *data* with libmagic; return it iff in allowlist.

    libmagic often classifies csv/md as text/plain. We promote based on
    extension only when the sniffed type is text/plain. Raises
    :class:`AttachmentRejected` if the resulting MIME is not allowed.
    """
    sniffed = magic.from_buffer(data, mime=True)

    if sniffed == "text/plain":
        ext = ""
        i = declared_filename.rfind(".")
        if i >= 0:
            ext = declared_filename[i:].lower()
        hinted = _EXT_HINTS.get(ext)
        if hinted:
            sniffed = hinted

    if sniffed not in ALLOWED_MIMES:
        raise AttachmentRejected(
            f"{declared_filename} ({sniffed}) is not a supported reference type. "
            "Supported: PDF, images (png/jpeg/gif/webp), Excel (xlsx), "
            "Word (docx), CSV, txt, md."
        )
    return sniffed


EXTRACTION_MAX_BYTES = 200_000  # 200 KB cap on per-file extracted text


def _truncate(text: str) -> str:
    if len(text.encode("utf-8")) <= EXTRACTION_MAX_BYTES:
        return text
    encoded = text.encode("utf-8")[:EXTRACTION_MAX_BYTES]
    truncated = encoded.decode("utf-8", errors="ignore")
    remaining = len(text.encode("utf-8")) - len(encoded)
    return truncated + f"\n\n[truncated — {remaining} more bytes]"


def _xlsx_to_markdown(data: bytes) -> str:
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    out: list[str] = []
    for ws in wb.worksheets:
        out.append(f"### Sheet: {ws.title}\n")
        first_row = True
        for row in ws.iter_rows(values_only=True):
            if all(c is None for c in row):
                continue
            cells = ["" if c is None else str(c) for c in row]
            out.append("| " + " | ".join(cells) + " |")
            if first_row:
                out.append("| " + " | ".join("---" for _ in cells) + " |")
                first_row = False
        out.append("")
    return "\n".join(out)


def _csv_to_text(data: bytes) -> str:
    return data.decode("utf-8", errors="replace")


def _docx_to_text(data: bytes) -> str:
    from docx import Document  # local import: python-docx is a heavier dependency
    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text)


_EXTRACTORS = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": _xlsx_to_markdown,
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": _docx_to_text,
    "text/csv": _csv_to_text,
    "text/markdown": _csv_to_text,  # pass-through; markdown is already text
    # text/plain is also passthrough but Claude takes it natively as a document block.
}


def extract_text(data: bytes, *, mime: str) -> str | None:
    """Return extracted text for non-native types, or None for native types.

    Output is capped at EXTRACTION_MAX_BYTES; oversize results get a
    ``[truncated — N more bytes]`` suffix.
    """
    extractor = _EXTRACTORS.get(mime)
    if extractor is None:
        return None
    raw = extractor(data)
    return _truncate(raw)


# ---------------------------------------------------------------------------
# Storage service — disk I/O + DB row creation
# ---------------------------------------------------------------------------

MAX_FILE_BYTES = 10 * 1024 * 1024   # 10 MB per file
MAX_SESSION_BYTES = 40 * 1024 * 1024  # 40 MB per session


def _upload_root() -> Path:
    root = Path(_settings.vision_upload_dir)
    root.mkdir(parents=True, exist_ok=True)
    return root


def _session_dir(session_id: str) -> Path:
    d = _upload_root() / session_id
    d.mkdir(parents=True, exist_ok=True)
    return d


def _unique_in_session(db_filenames: set[str], filename: str) -> str:
    if filename not in db_filenames:
        return filename
    root, ext = filename.rsplit(".", 1) if "." in filename else (filename, "")
    suffix_ext = f".{ext}" if ext else ""
    for n in range(2, 1000):
        candidate = f"{root}-{n}{suffix_ext}"
        if candidate not in db_filenames:
            return candidate
    raise AttachmentRejected("too many duplicate filenames in session")


async def store_attachment(
    db: AsyncSession,
    *,
    session_id: str,
    raw: bytes,
    declared_filename: str,
) -> VisionChatAttachment:
    """Validate, sniff, extract, write to disk, insert row. Commit is caller's job."""
    # Size check FIRST — before hitting libmagic or disk.
    if len(raw) > MAX_FILE_BYTES:
        raise AttachmentRejected(
            f"{declared_filename} is {len(raw) // (1024 * 1024)} MB — max 10 MB per file."
        )

    result = await db.execute(
        select(func.coalesce(func.sum(VisionChatAttachment.size_bytes), 0))
        .where(VisionChatAttachment.session_id == session_id)
    )
    current_total = int(result.scalar_one())
    if current_total + len(raw) > MAX_SESSION_BYTES:
        raise AttachmentRejected(
            "Adding this would exceed the 40 MB per-session attachment limit."
        )

    safe_name = sanitize_filename(declared_filename)
    mime = sniff_and_validate_mime(raw, declared_filename=safe_name)

    existing = await db.execute(
        select(VisionChatAttachment.filename).where(
            VisionChatAttachment.session_id == session_id
        )
    )
    used = {row[0] for row in existing.all()}
    final_name = _unique_in_session(used, safe_name)

    disk_id = _uuid.uuid4().hex
    disk_path = _session_dir(session_id) / f"{disk_id}-{final_name}"
    disk_path.write_bytes(raw)

    extracted = extract_text(raw, mime=mime)

    att = VisionChatAttachment(
        id=str(_uuid.uuid4()),
        session_id=session_id,
        filename=final_name,
        mime_type=mime,
        size_bytes=len(raw),
        disk_path=str(disk_path),
        extracted_text=extracted,
    )
    db.add(att)
    return att


async def delete_attachment(db: AsyncSession, *, attachment_id: str) -> None:
    """Delete an attachment IFF it has not yet been sent in a chat turn."""
    att = await db.get(VisionChatAttachment, attachment_id)
    if att is None:
        raise AttachmentRejected("attachment not found")
    if att.sent_at is not None:
        raise AttachmentRejected("cannot delete an attachment already sent in a chat turn")
    path = Path(att.disk_path)
    if path.exists():
        try:
            path.unlink()
        except OSError:
            pass
    await db.delete(att)


def cleanup_session_dir(session_id: str) -> None:
    """Remove the session's upload dir on disk (no-op if absent or unreadable).

    Reads the upload root directly rather than calling ``_upload_root()`` so
    cleanup never accidentally creates the parent directory — important when
    invoked from environments (tests, restricted users) that can't write to
    the default ``/var/lib/...`` path.
    """
    d = Path(_settings.vision_upload_dir) / session_id
    if d.exists():
        shutil.rmtree(d, ignore_errors=True)


# ---------------------------------------------------------------------------
# Multi-block message builder
# ---------------------------------------------------------------------------

_IMAGE_MIMES = {"image/png", "image/jpeg", "image/gif", "image/webp"}


async def build_chat_blocks(
    db: AsyncSession,
    *,
    user_text: str,
    attachment_ids: list[str] | None,
) -> list[dict]:
    """Assemble a multi-block user message: text + attachments.

    Side effect: marks the included attachments as ``sent_at = now()``.
    Caller is responsible for committing.
    """
    blocks: list[dict] = [{"type": "text", "text": user_text}]
    if not attachment_ids:
        return blocks

    now = datetime.now(timezone.utc)
    for aid in attachment_ids:
        att = await db.get(VisionChatAttachment, aid)
        if att is None:
            continue
        raw = Path(att.disk_path).read_bytes()
        b64 = base64.b64encode(raw).decode("ascii")

        if att.mime_type == "application/pdf":
            blocks.append({
                "type": "document",
                "source": {"type": "base64", "media_type": "application/pdf", "data": b64},
                "title": att.filename,
            })
        elif att.mime_type == "text/plain":
            blocks.append({
                "type": "document",
                "source": {
                    "type": "text",
                    "media_type": "text/plain",
                    "data": raw.decode("utf-8", errors="replace"),
                },
                "title": att.filename,
            })
        elif att.mime_type in _IMAGE_MIMES:
            blocks.append({
                "type": "image",
                "source": {"type": "base64", "media_type": att.mime_type, "data": b64},
            })
        else:
            prefix = f"--- Attached file: {att.filename} ({att.mime_type}) ---\n"
            body = att.extracted_text or "(no extractable text)"
            blocks.append({"type": "text", "text": prefix + body})

        att.sent_at = now

    return blocks
